import openpyxl

def get_test_data(test_case_name, file_path, sheet_name):
    book = openpyxl.load_workbook(file_path)
    sheet = book[sheet_name]
    data = []

    for i in range(1, sheet.max_row + 1):
        dic = {}
        if sheet.cell(row=i, column=1).value == test_case_name:
            for j in range(2, sheet.max_column + 1):
                dic[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
            data.append(dic)
    return data

# def update_excel(file_path, search_term, sheet_name):
